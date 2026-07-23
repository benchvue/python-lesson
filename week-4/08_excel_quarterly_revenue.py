"""Advanced: 5 years of quarterly revenue in Excel.

Builds a 5 x 4 grid (2022-2026 x Q1-Q4), then adds
live Excel formulas for row totals, quarter averages and YoY growth.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

YEARS = [2022, 2023, 2024, 2025, 2026]
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

revenue = {
    2022: [120000, 135000, 128000, 150000],
    2023: [140000, 152000, 149000, 171000],
    2024: [165000, 178000, 172000, 199000],
    2025: [190000, 205000, 201000, 232000],
    2026: [215000, 236000, 228000, 264000],
}

wb = Workbook()
ws = wb.active
ws.title = "Quarterly Revenue"

# --- header row ---
headers = ["Year"] + QUARTERS + ["Total", "Growth %"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

# --- one row per year, with live formulas ---
for i, year in enumerate(YEARS):
    r = i + 2                       # data starts on row 2
    ws.cell(row=r, column=1, value=year)
    for q in range(4):
        ws.cell(row=r, column=q + 2, value=revenue[year][q])
    # Total = SUM(B:E) for this row
    ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
    # Growth % vs the previous year (first year has no prior)
    if i > 0:
        ws.cell(row=r, column=7, value=f"=(F{r}-F{r-1})/F{r-1}")
        ws.cell(row=r, column=7).number_format = "0.0%"

# --- footer: average per quarter across the 5 years ---
last = len(YEARS) + 1
ws.cell(row=last + 1, column=1, value="Avg").font = Font(bold=True)
for q in range(4):
    col = get_column_letter(q + 2)
    ws.cell(row=last + 1, column=q + 2,
            value=f"=AVERAGE({col}2:{col}{last})")

# --- formatting ---
for r in range(2, last + 2):
    for c in range(2, 7):
        ws.cell(row=r, column=c).number_format = "#,##0"
ws.column_dimensions["A"].width = 10
for c in range(2, 8):
    ws.column_dimensions[get_column_letter(c)].width = 12
ws.freeze_panes = "B2"

wb.save("quarterly_revenue.xlsx")
print("quarterly_revenue.xlsx created (5 years x 4 quarters).")

# --- same math in Python, so you can check the numbers ---
print(f"\n{'Year':<8}{'Total':>12}{'Growth':>10}")
prev = None
for year in YEARS:
    total = sum(revenue[year])
    growth = "" if prev is None else f"{(total - prev) / prev:.1%}"
    print(f"{year:<8}{total:>12,}{growth:>10}")
    prev = total
