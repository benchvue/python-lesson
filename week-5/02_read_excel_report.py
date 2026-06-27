from openpyxl import load_workbook

workbook = load_workbook("income_statement.xlsx")
sheet = workbook["Income Statement"]

for row in sheet.iter_rows(values_only=True):
    print(row)
